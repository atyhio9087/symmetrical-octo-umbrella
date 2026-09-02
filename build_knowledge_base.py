"""
Offline knowledge-base compaction: reads a CSV/Excel of case studies and precomputes everything
the matching worker needs at request time — sentence embeddings for semantic search, and the
document text used for BM25 keyword search — into a single shipped artifact.

Run this whenever the case-study database changes, then redeploy:

    python build_knowledge_base.py --input assets/knowledge_base.csv

Output (both required by match_projects.py):
    assets/knowledge_base_vectordb.npy   float32 embeddings, one row per project
    assets/knowledge_base_vectordb.json  project fields + document text + build metadata,
                                          in the same row order as the .npy file

Why precompute instead of embedding at request time: this app's live-matching UI calls the match
endpoint on every stakeholder-field edit, but the case-study database itself changes rarely (an
admin update, not a per-request thing). Embedding all case studies fresh on every server boot (or
worse, per request) burns time and model-inference calls for work whose result almost never
changes. Baking it into a static artifact means the running app only ever has to embed the one
new query string per request.
"""

import argparse
import csv
import io
import json
import re
import sys
from datetime import datetime, timezone

import numpy as np

EMBEDDING_MODEL_NAME = "sentence-transformers/all-MiniLM-L6-v2"

# Column aliases, matching the fuzzy header-matching used by the app's own CSV/Excel upload path
# (src/utils/projectNormalizer.ts) — so the same file you'd otherwise upload works here unchanged.
FIELD_ALIASES = {
    "sNo": ["sno", "s.no", "serial"],
    "insightPeriod": ["insight period", "period"],
    "client": ["client", "company"],
    "projectType": ["project type", "type"],
    "businessArea": ["business area", "area"],
    "technologyUsed": ["technology used", "tech", "technology"],
    "deliverableName": ["deliverable name", "deliverable", "project name"],
    "problemStatement": ["problem statement", "business problem", "problem"],
    "objective": ["objective", "goal"],
    "approach": ["approach", "methodology", "solution"],
    "impactCreated": ["impact created", "impact details"],
    "impactType": ["impact type"],
    "valueImpact": ["value impact", "value", "impact value"],
}


def _normalize_header(h: str) -> str:
    return re.sub(r"[^a-z0-9]", "", h.lower())


def normalize_row(raw_row: dict, row_id: str) -> dict:
    normalized_keys = {_normalize_header(k): k for k in raw_row.keys()}

    def find_value(aliases):
        for alias in aliases:
            alias_norm = _normalize_header(alias)
            for norm_key, original_key in normalized_keys.items():
                if alias_norm in norm_key:
                    val = raw_row.get(original_key)
                    if val is not None and str(val).strip() != "":
                        return str(val).strip()
        return ""

    fields = {name: find_value(aliases) for name, aliases in FIELD_ALIASES.items()}
    fields["id"] = row_id
    fields["deliverableName"] = fields["deliverableName"] or "Unnamed Project"
    fields["impactCreated"] = fields["impactCreated"] or "N/A"
    fields["impactType"] = fields["impactType"] or "Operational"
    return fields


def read_rows(path: str) -> list:
    if path.lower().endswith((".xlsx", ".xls")):
        try:
            import openpyxl
        except ImportError:
            sys.exit(
                "Reading .xlsx requires openpyxl: pip install openpyxl (or export the sheet as .csv instead)."
            )
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows_iter = ws.iter_rows(values_only=True)
        headers = [str(h) if h is not None else "" for h in next(rows_iter)]
        raw_rows = []
        for row in rows_iter:
            if all(cell is None for cell in row):
                continue
            raw_rows.append({headers[i]: row[i] for i in range(len(headers)) if i < len(row)})
        return raw_rows
    else:
        with io.open(path, "r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            return [row for row in reader if any((v or "").strip() for v in row.values())]


def build_document(p: dict) -> str:
    """The text that gets embedded and BM25-indexed for each project. Deliberately includes more
    signal than just the name/area/client/value used previously — Problem Statement, Objective,
    Approach, Technology Used, and Impact Type all carry real semantic content that helps both
    embedding similarity and BM25 keyword matches, and are skipped cleanly when blank rather than
    injecting empty/placeholder text."""
    parts = []
    if p["deliverableName"]:
        parts.append(f"Deliverable: {p['deliverableName']}.")
    if p["businessArea"]:
        parts.append(f"Business Area: {p['businessArea']}.")
    if p["projectType"]:
        parts.append(f"Project Type: {p['projectType']}.")
    if p["technologyUsed"]:
        parts.append(f"Technology Used: {p['technologyUsed']}.")
    if p["client"]:
        parts.append(f"Client: {p['client']}.")
    if p["problemStatement"]:
        parts.append(f"Problem: {p['problemStatement']}")
    if p["objective"]:
        parts.append(f"Objective: {p['objective']}")
    if p["approach"]:
        parts.append(f"Approach: {p['approach']}")
    if p["impactType"]:
        parts.append(f"Impact Type: {p['impactType']}.")
    if p["valueImpact"]:
        parts.append(f"Value (Impact): {p['valueImpact']}.")
    # A dedicated keyword-dense tail (business area / tech / project type / impact type, repeated
    # verbatim) gives BM25's exact-term matching something concentrated to latch onto, separate
    # from the more narrative sentence-style fields above that primarily help the embedding model.
    keywords = ", ".join(
        x for x in [p["businessArea"], p["projectType"], p["technologyUsed"], p["impactType"]] if x
    )
    if keywords:
        parts.append(f"Keywords: {keywords}.")
    return " ".join(parts)


def main():
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--input", default="assets/knowledge_base.csv", help="Path to the CSV/Excel knowledge base file")
    parser.add_argument("--output-prefix", default="assets/knowledge_base_vectordb", help="Output path prefix (writes <prefix>.npy and <prefix>.json)")
    args = parser.parse_args()

    print(f"Reading {args.input} ...")
    raw_rows = read_rows(args.input)
    if not raw_rows:
        sys.exit(f"No rows found in {args.input}. Nothing to build.")

    projects = [normalize_row(row, f"kb-{idx}") for idx, row in enumerate(raw_rows)]
    documents = [build_document(p) for p in projects]

    print(f"Loading embedding model ({EMBEDDING_MODEL_NAME}) — this can take a while on first run...")
    from sentence_transformers import SentenceTransformer

    model = SentenceTransformer(EMBEDDING_MODEL_NAME)

    print(f"Embedding {len(documents)} case studies...")
    embeddings = model.encode(documents, convert_to_numpy=True, show_progress_bar=True).astype(np.float32)

    npy_path = f"{args.output_prefix}.npy"
    json_path = f"{args.output_prefix}.json"

    np.save(npy_path, embeddings)
    with io.open(json_path, "w", encoding="utf-8") as f:
        json.dump(
            {
                "model": EMBEDDING_MODEL_NAME,
                "builtAt": datetime.now(timezone.utc).isoformat(),
                "count": len(projects),
                "projects": projects,
                "documents": documents,
            },
            f,
            indent=2,
        )

    print(f"Wrote {npy_path} ({embeddings.shape[0]} x {embeddings.shape[1]}) and {json_path}.")
    print("Redeploy the app with these two files under assets/ to pick up the update.")


if __name__ == "__main__":
    main()
